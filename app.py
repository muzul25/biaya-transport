import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

st.set_page_config(page_title="Generator Template SPD", page_icon="📄")

st.title("📄 Generator Template SPD")

# Fungsi buat PDF sederhana dari data
def buat_pdf(nama, data_row):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 50, f"Surat Perjalanan Dinas - {nama}")

    c.setFont("Helvetica", 12)
    y = height - 100
    c.drawString(50, y, f"Nama: {data_row['Nama']}")
    y -= 20
    c.drawString(50, y, f"Transport Pesawat: Rp {data_row['Transport Pesawat']:,}")
    y -= 20
    c.drawString(50, y, f"Transport Kereta: Rp {data_row['Transport Kereta']:,}")
    y -= 20
    c.drawString(50, y, f"Transport Bandara/Stasiun Asal: Rp {data_row['Transport Bandara/Stasiun Asal']:,}")
    y -= 20
    c.drawString(50, y, f"Transport Bandara/Stasiun Hotel: Rp {data_row['Transport Bandara/Stasiun Hotel']:,}")
    y -= 20
    c.drawString(50, y, f"Transport Lokal: Rp {data_row['Transport Lokal']:,}")
    y -= 30

    total = sum([
        data_row["Transport Pesawat"],
        data_row["Transport Kereta"],
        data_row["Transport Bandara/Stasiun Asal"],
        data_row["Transport Bandara/Stasiun Hotel"],
        data_row["Transport Lokal"]
    ])
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, f"TOTAL: Rp {total:,}")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer

# Upload database dan template
db_file = st.file_uploader("Upload Database (Excel/CSV)", type=["xlsx", "csv"])
template_file = st.file_uploader("Upload Template SPD (Excel)", type=["xlsx"])

if db_file and template_file:
    # Baca database
    if db_file.name.endswith(".csv"):
        df = pd.read_csv(db_file)
    else:
        df = pd.read_excel(db_file)

    required_cols = [
        "Nama",
        "Transport Pesawat",
        "Transport Kereta",
        "Transport Bandara/Stasiun Asal",
        "Transport Bandara/Stasiun Hotel",
        "Transport Lokal"
    ]

    if not all(col in df.columns for col in required_cols):
        st.error(f"Kolom database harus mengandung: {', '.join(required_cols)}")
    else:
        st.subheader("Preview Database")
        st.dataframe(df.head())

        nama_terpilih = st.selectbox("Pilih Nama", df["Nama"].unique())

        # Generate Excel per nama
        if st.button("🔄 Generate Excel & PDF (Satu Nama)"):
            data_row = df[df["Nama"] == nama_terpilih].iloc[0]

            # === Excel ===
            wb = load_workbook(template_file)
            ws = wb.active
            ws["D26"] = data_row["Nama"]
            ws["C11"] = data_row["Transport Pesawat"]
            ws["C12"] = data_row["Transport Kereta"]
            ws["C13"] = data_row["Transport Bandara/Stasiun Asal"]
            ws["C14"] = data_row["Transport Bandara/Stasiun Hotel"]
            ws["C15"] = data_row["Transport Lokal"]
            ws["C17"] = sum([
                data_row["Transport Pesawat"],
                data_row["Transport Kereta"],
                data_row["Transport Bandara/Stasiun Asal"],
                data_row["Transport Bandara/Stasiun Hotel"],
                data_row["Transport Lokal"]
            ])

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # === PDF ===
            pdf_buffer = buat_pdf(nama_terpilih, data_row)

            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_buffer,
                    file_name=f"Template_{nama_terpilih}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col2:
                st.download_button(
                    label="📄 Download PDF",
                    data=pdf_buffer,
                    file_name=f"Template_{nama_terpilih}.pdf",
                    mime="application/pdf"
                )

        # Generate semua (ZIP)
        if st.button("📦 Generate Semua Template (Excel + PDF ZIP)"):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for _, row in df.iterrows():
                    # === Excel ===
                    wb = load_workbook(template_file)
                    ws = wb.active
                    ws["D26"] = row["Nama"]
                    ws["C11"] = row["Transport Pesawat"]
                    ws["C12"] = row["Transport Kereta"]
                    ws["C13"] = row["Transport Bandara/Stasiun Asal"]
                    ws["C14"] = row["Transport Bandara/Stasiun Hotel"]
                    ws["C15"] = row["Transport Lokal"]
                    ws["C17"] = sum([
                        row["Transport Pesawat"],
                        row["Transport Kereta"],
                        row["Transport Bandara/Stasiun Asal"],
                        row["Transport Bandara/Stasiun Hotel"],
                        row["Transport Lokal"]
                    ])

                    excel_buf = io.BytesIO()
                    wb.save(excel_buf)
                    excel_buf.seek(0)
                    zf.writestr(f"Template_{row['Nama']}.xlsx", excel_buf.read())

                    # === PDF ===
                    pdf_buf = buat_pdf(row["Nama"], row)
                    zf.writestr(f"Template_{row['Nama']}.pdf", pdf_buf.read())

            zip_buffer.seek(0)
            st.download_button(
                label="📦 Download ZIP Semua",
                data=zip_buffer,
                file_name="Semua_Template_SPD.zip",
                mime="application/zip"
            )
